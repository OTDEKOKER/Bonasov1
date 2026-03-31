import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from import_selected_q3_workbook import (
    canonical_indicator_key,
    get_indicator_resolution_priority,
    parse_sheet,
    resolve_indicator,
)


WORKBOOK_PATH = Path(
    r"C:\Users\dekok\Downloads\CSOs Reports\CSOs Reports\TEBELOPELE\TEBELOPELE- NAHPA REPORTING TEMPLATE 2025_26 Q3.xlsm"
)

PROJECT_ID = 2
PERIOD_START = "2025-10-01"
PERIOD_END = "2025-12-31"

REPORT_PATH = Path("reports/tebelopele-resync-audit.json")
SQL_PATH = Path("reports/tebelopele-resync.sql")

SHEET_TO_ORG = {
    "TEBELOPELE": "TEBELOPELE",
    "SENTEBALE": "Sentebale",
    "BOFWA": "BOFWA",
    "Stepping Stone International": "Stepping Stone International",
    "Bobonong Home Based Care": "Bobonong Home Based Care",
    "Mabogo aa Thebana Association S": "Mabogo aa Thebana Association South",
    "Ovajua": "Ovajhu",
    "Gumare Advisory": "Gumare Advisory",
    "Positive Moments": "Positive Moments",
    "Mind Power": "Mind Power",
    "Mopipi International Trust": "Mopipi International Trust",
    "House of Angels": "House of Angels",
    "Inspired Hozirons": "Inspired Horizons",
    "Hope Worldwide": "Hope Worldwide",
    "Rena le Seabe": "Rena le Seabe",
    "Social Dialogue": "Social Dialogue",
    "INERELA": "INERELA",
}


class IndicatorProxy:
    def __init__(self, indicator_id, code, name, is_active):
        self.id = int(indicator_id)
        self.code = code
        self.name = name
        self.is_active = bool(is_active)


def run_psql(query: str) -> str:
    command = [
        "docker",
        "exec",
        "bonaso-local-stack-postgres-1",
        "psql",
        "-U",
        "bonaso",
        "-d",
        "bonaso",
        "-F",
        "\t",
        "-A",
        "-c",
        query,
    ]
    return subprocess.check_output(command, text=True)


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def load_indicators():
    query = """
select id, code, name, is_active
from indicators_indicator
order by id;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    indicator_by_key = {}
    for line in lines:
        parts = line.split("\t", 3)
        if len(parts) != 4:
            continue
        if parts[0] == "id":
            continue
        candidate = IndicatorProxy(parts[0], parts[1], parts[2], parts[3] == "t")
        key = canonical_indicator_key(candidate.name)
        existing = indicator_by_key.get(key)
        if existing is None or get_indicator_resolution_priority(candidate) < get_indicator_resolution_priority(existing):
            indicator_by_key[key] = candidate
    return indicator_by_key


def load_organizations():
    query = """
select id, name
from organizations_organization
where id = 1 or parent_id = 1
order by name;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    result = {}
    for line in lines:
        if line.startswith("id\t"):
            continue
        org_id, name = line.split("\t", 1)
        result[name] = int(org_id)
    return result


def load_existing_rows():
    query = f"""
with recursive orgs as (
  select id,name
  from organizations_organization
  where id = 1
  union all
  select o.id,o.name
  from organizations_organization o
  join orgs on o.parent_id = orgs.id
)
select
  a.id,
  o.name as organization_name,
  a.organization_id,
  a.indicator_id,
  i.code,
  i.name,
  a.value::text
from aggregates_aggregate a
join orgs o on o.id = a.organization_id
join indicators_indicator i on i.id = a.indicator_id
where a.project_id = {PROJECT_ID}
  and a.period_start = '{PERIOD_START}'
  and a.period_end = '{PERIOD_END}'
order by o.name, i.name;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    parsed = []
    for row in lines:
        if row.startswith("id\t"):
            continue
        aggregate_id, org_name, organization_id, indicator_id, code, name, value = row.split("\t", 6)
        parsed.append(
            {
                "aggregate_id": int(aggregate_id),
                "organization_name": org_name,
                "organization_id": int(organization_id),
                "indicator_id": int(indicator_id),
                "code": code,
                "name": name,
                "value": json.loads(value),
            }
        )
    return parsed


def main():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    indicator_candidates = load_indicators()
    organizations = load_organizations()
    existing_rows = load_existing_rows()
    existing_by_org_indicator = {
        (row["organization_name"], row["indicator_id"]): row
        for row in existing_rows
    }

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)

    updates = []
    inserts = []
    missing = []
    desired_keys = set()

    for sheet_name, org_name in SHEET_TO_ORG.items():
        if sheet_name not in workbook.sheetnames:
            missing.append({"sheet": sheet_name, "organization": org_name, "reason": "sheet missing"})
            continue

        org_id = organizations.get(org_name)
        if org_id is None:
            missing.append({"sheet": sheet_name, "organization": org_name, "reason": "organization missing"})
            continue

        items = parse_sheet(workbook[sheet_name])
        for item in items:
            indicator = resolve_indicator(item["title"], indicator_candidates, code=item["code"])
            if indicator is None:
                missing.append(
                    {
                        "sheet": sheet_name,
                        "organization": org_name,
                        "code": item["code"],
                        "title": item["title"],
                        "reason": "indicator missing",
                    }
                )
                continue

            desired_keys.add((org_name, indicator.id))
            existing = existing_by_org_indicator.get((org_name, indicator.id))
            if existing is None:
                inserts.append(
                    {
                        "organization_name": org_name,
                        "organization_id": org_id,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "value": item["value"],
                    }
                )
                continue

            if existing["value"] != item["value"]:
                updates.append(
                    {
                        "aggregate_id": existing["aggregate_id"],
                        "organization_name": org_name,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "before": existing["value"],
                        "after": item["value"],
                    }
                )

    stale_rows = [
        row
        for row in existing_rows
        if (row["organization_name"], row["indicator_id"]) not in desired_keys
    ]

    sql_lines = ["begin;"]
    for update in updates:
        value_json = json.dumps(update["after"], ensure_ascii=True)
        sql_lines.append(
            "update aggregates_aggregate "
            f"set value = {sql_literal(value_json)}::jsonb, "
            "status = 'approved', "
            "reviewed_at = now(), "
            "notes = coalesce(notes,'') || ' | resynced from Tebelopele workbook' "
            f"where id = {update['aggregate_id']};"
        )

    for insert in inserts:
        value_json = json.dumps(insert["value"], ensure_ascii=True)
        notes = (
            f"Imported from {WORKBOOK_PATH.name} | sheet={insert['organization_name']} | "
            f"code={insert['code']} | resynced from Tebelopele workbook"
        )
        sql_lines.append(
            "insert into aggregates_aggregate "
            "(period_start, period_end, value, notes, created_at, updated_at, created_by_id, "
            "indicator_id, organization_id, project_id, reviewed_at, reviewed_by_id, status) values ("
            f"{sql_literal(PERIOD_START)}, "
            f"{sql_literal(PERIOD_END)}, "
            f"{sql_literal(value_json)}::jsonb, "
            f"{sql_literal(notes)}, "
            "now(), now(), 5, "
            f"{insert['indicator_id']}, "
            f"{insert['organization_id']}, "
            f"{PROJECT_ID}, "
            "now(), 5, 'approved');"
        )

    for stale in stale_rows:
        sql_lines.append(
            "delete from aggregates_aggregatechangelog "
            f"where aggregate_id = {stale['aggregate_id']};"
        )
        sql_lines.append(f"delete from aggregates_aggregate where id = {stale['aggregate_id']};")
    sql_lines.append("commit;")

    SQL_PATH.write_text("\n".join(sql_lines), encoding="utf-8")

    report = {
        "workbook": str(WORKBOOK_PATH),
        "existing_rows": len(existing_rows),
        "updates": len(updates),
        "inserts": len(inserts),
        "stale_rows": len(stale_rows),
        "missing": missing,
        "updated_rows": updates,
        "inserted_rows": inserts,
        "stale_row_details": stale_rows,
        "sql_path": str(SQL_PATH),
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "updates": len(updates),
                "inserts": len(inserts),
                "stale_rows": len(stale_rows),
                "missing": len(missing),
            },
            indent=2,
        )
    )
    print(f"Report: {REPORT_PATH}")
    print(f"SQL: {SQL_PATH}")


if __name__ == "__main__":
    main()
