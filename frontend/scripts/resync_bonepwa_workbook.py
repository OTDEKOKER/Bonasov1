import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from import_selected_q3_workbook import (
    canonical_indicator_key,
    get_indicator_resolution_priority,
    parse_sheet,
    resolve_indicator,
)


WORKBOOK_PATH = Path(
    r"C:\Users\dekok\Downloads\Q3 REPORTS\Q3 REPORTS\Q3 BONEPWA+\BONEPWA Q3 reporting template .xlsx"
)

PROJECT_ID = 2
PERIOD_START = "2025-10-01"
PERIOD_END = "2025-12-31"

REPORT_PATH = Path("reports/bonepwa-resync-audit.json")
SQL_PATH = Path("reports/bonepwa-resync.sql")

SHEET_TO_ORG = {
    "BONEPWA+": "BONEPWA",
    "Healing Hearts": "Healing Hearts",
    "Sunshine": "Sunshine",
    "CEYOHO": "CEYOHO",
    "Thusang Bana": "Thusang Bana",
    "Thabologo Support Group": "Thhabologo Support Group",
    "Matlhogonolo Charitable Society": "Mathogonolo Charitable Society",
    "Kebotlhokwa": "Kebotlhekwa",
    "Marang": "Marang",
    "Kasane": "Kasane",
    "Tutume AIDS Fighters": "Tutume AIDS Fighters",
    "Lapologang": "Lapologang",
    "Leitlho La Sechaba": "Leitho la Sechaba",
    "Tozwimilidizha Muti Amuchile ": "Tozwimilidzha Muti Amuchile Support Group",
    "BCNS": "BCNS",
    "Baikamogedi Support Group": "Baikamogedi Support Group/POT",
    "Ditsheganwe Support Group": "Ditshegwane Support Group",
    "Motheo Support Group ": "Motheo Support Group",
    "A he eme": "A he eme",
    "Ghantsi Support Group": "Ghantsi Support Group",
    "Jwaneng Support Group": "Jwaneng Support Group",
    "Incredible Minds": "Incredible Minds",
    "Guardian Angels Orphans Society": "Guardian Angel Orphans Society",
    "Old Naledi Support Group": "Old Naledi Support Group",
    "BORNUS": "BORNUS",
    "Mafolofolo Support Group": "Omaweneno Support Group",
    "Cynthia's Child Care Trust": "Cynthia's Child Care Trust",
    "Botswana Bailor Childrens": "Botswana Bailor Childrens",
    "Center for TB": "Center for TB",
    "Diphalana": "Diphalana",
}


class IndicatorProxy:
    def __init__(self, indicator_id, code, name, is_active):
        self.id = int(indicator_id)
        self.code = code
        self.name = name
        self.is_active = bool(is_active)


def run_psql(query: str) -> str:
    command = [
        "docker",
        "exec",
        "bonaso-local-stack-postgres-1",
        "psql",
        "-U",
        "bonaso",
        "-d",
        "bonaso",
        "-F",
        "\t",
        "-A",
        "-c",
        query,
    ]
    return subprocess.check_output(command, text=True)


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def load_indicators():
    query = """
select id, code, name, is_active
from indicators_indicator
order by id;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    indicator_by_key = {}
    for line in lines:
        parts = line.split("\t", 3)
        if len(parts) != 4 or parts[0] == "id":
            continue
        candidate = IndicatorProxy(parts[0], parts[1], parts[2], parts[3] == "t")
        key = canonical_indicator_key(candidate.name)
        existing = indicator_by_key.get(key)
        if existing is None or get_indicator_resolution_priority(candidate) < get_indicator_resolution_priority(existing):
            indicator_by_key[key] = candidate
    return indicator_by_key


def load_organizations():
    query = """
select id, name
from organizations_organization
where id = 112 or parent_id = 112
order by name;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    result = {}
    for line in lines:
        if line.startswith("id\t"):
            continue
        org_id, name = line.split("\t", 1)
        result[name] = int(org_id)
    return result


def load_existing_rows():
    query = f"""
select
  a.id,
  o.name as organization_name,
  a.organization_id,
  a.indicator_id,
  i.code,
  i.name,
  a.value::text
from aggregates_aggregate a
join organizations_organization o on o.id = a.organization_id
join indicators_indicator i on i.id = a.indicator_id
where a.project_id = {PROJECT_ID}
  and a.period_start = '{PERIOD_START}'
  and a.period_end = '{PERIOD_END}'
  and (a.organization_id = 112 or o.parent_id = 112)
order by o.name, i.name;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    parsed = []
    for row in lines:
        if row.startswith("id\t"):
            continue
        aggregate_id, org_name, organization_id, indicator_id, code, name, value = row.split("\t", 6)
        parsed.append(
            {
                "aggregate_id": int(aggregate_id),
                "organization_name": org_name,
                "organization_id": int(organization_id),
                "indicator_id": int(indicator_id),
                "code": code,
                "name": name,
                "value": json.loads(value),
            }
        )
    return parsed


def main():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    indicator_candidates = load_indicators()
    organizations = load_organizations()
    existing_rows = load_existing_rows()
    existing_by_org_indicator = {
        (row["organization_name"], row["indicator_id"]): row
        for row in existing_rows
    }

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)

    updates = []
    inserts = []
    missing = []
    desired_keys = set()

    for sheet_name, org_name in SHEET_TO_ORG.items():
        if sheet_name not in workbook.sheetnames:
            missing.append({"sheet": sheet_name, "organization": org_name, "reason": "sheet missing"})
            continue

        org_id = organizations.get(org_name)
        if org_id is None:
            missing.append({"sheet": sheet_name, "organization": org_name, "reason": "organization missing"})
            continue

        items = parse_sheet(workbook[sheet_name])
        for item in items:
            indicator = resolve_indicator(item["title"], indicator_candidates, code=item["code"])
            if indicator is None:
                missing.append(
                    {
                        "sheet": sheet_name,
                        "organization": org_name,
                        "code": item["code"],
                        "title": item["title"],
                        "reason": "indicator missing",
                    }
                )
                continue

            desired_keys.add((org_name, indicator.id))
            existing = existing_by_org_indicator.get((org_name, indicator.id))
            if existing is None:
                inserts.append(
                    {
                        "sheet_name": sheet_name,
                        "organization_name": org_name,
                        "organization_id": org_id,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "value": item["value"],
                    }
                )
                continue

            if existing["value"] != item["value"]:
                updates.append(
                    {
                        "aggregate_id": existing["aggregate_id"],
                        "sheet_name": sheet_name,
                        "organization_name": org_name,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "before": existing["value"],
                        "after": item["value"],
                    }
                )

    stale_rows = [
        row
        for row in existing_rows
        if (row["organization_name"], row["indicator_id"]) not in desired_keys
    ]

    sql_lines = ["begin;"]
    for update in updates:
        value_json = json.dumps(update["after"], ensure_ascii=True)
        sql_lines.append(
            "update aggregates_aggregate "
            f"set value = {sql_literal(value_json)}::jsonb, "
            "status = 'approved', "
            "reviewed_at = now(), "
            "notes = coalesce(notes,'') || ' | resynced from BONEPWA Q3 workbook' "
            f"where id = {update['aggregate_id']};"
        )

    for insert in inserts:
        value_json = json.dumps(insert["value"], ensure_ascii=True)
        notes = (
            f"Imported from {WORKBOOK_PATH.name} | sheet={insert['sheet_name']} | "
            f"code={insert['code']} | resynced from BONEPWA Q3 workbook"
        )
        sql_lines.append(
            "insert into aggregates_aggregate "
            "(period_start, period_end, value, notes, created_at, updated_at, created_by_id, "
            "indicator_id, organization_id, project_id, reviewed_at, reviewed_by_id, status) "
            "values ("
            f"'{PERIOD_START}', '{PERIOD_END}', {sql_literal(value_json)}::jsonb, {sql_literal(notes)}, "
            "now(), now(), 5, "
            f"{insert['indicator_id']}, {insert['organization_id']}, {PROJECT_ID}, now(), 5, 'approved'"
            ");"
        )

    for stale in stale_rows:
        sql_lines.append(
            f"delete from aggregates_aggregatechangelog where aggregate_id = {stale['aggregate_id']};"
        )
        sql_lines.append(
            f"delete from aggregates_aggregate where id = {stale['aggregate_id']};"
        )
    sql_lines.append("commit;")

    REPORT_PATH.write_text(
        json.dumps(
            {
                "workbook": str(WORKBOOK_PATH),
                "updates": updates,
                "inserts": inserts,
                "stale_rows": stale_rows,
                "missing": missing,
                "summary": {
                    "updates": len(updates),
                    "inserts": len(inserts),
                    "stale_rows": len(stale_rows),
                    "missing": len(missing),
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    SQL_PATH.write_text("\n".join(sql_lines) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "report": str(REPORT_PATH),
                "sql": str(SQL_PATH),
                "updates": len(updates),
                "inserts": len(inserts),
                "stale_rows": len(stale_rows),
                "missing": len(missing),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
