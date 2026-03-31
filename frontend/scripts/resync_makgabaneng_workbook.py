import json
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from import_selected_q3_workbook import (
    canonical_indicator_key,
    get_indicator_resolution_priority,
    resolve_indicator,
)


WORKBOOK_PATH = Path(
    r"C:\Users\dekok\AppData\Local\Temp\8b85d847-fc10-4529-9c38-37db84548c1c_Q3 REPORTS (1).zip.Q3 REPORTS (1).zip\Q3 REPORTS\Q3 Makgabaneng\Quarter 3 2025 Makgabaneng NCD  REPORT .xlsx"
)

SHEET_TO_ORG = {
    "Masego Mental Health Org.": "Masego Mental Health",
    "Mabogo aa Thebana Association": "MAATA",
    "Bona Naledi": "BONA NALEDI",
    "HPP": "HPP",
    "APSA": "APSA",
    "Chobe Arts": "Chobe Arts",
    "NCONGO": "NCONGO",
    "Stop Smoking Support Group": "SSSG",
    "The Fighters Support Group": "TFSG",
    "Ultimate Youth": "Ultimate Youth",
    "ATN": "ATN",
    "Keitsholofetse": "Keitsholofetse",
    "VMF Valour": "Valour Mental Health",
    "Home of Hope": "Home of Hope",
    "The Just Hope Foundation": "Just Hope Foundation",
    "BONMEH": "BONMEH",
    "BOSASNet": "BOSASNet",
    "JOH": "JOH",
}

PROJECT_ID = 2
PERIOD_START = "2025-10-01"
PERIOD_END = "2025-12-31"

REPORT_PATH = Path("reports/makgabaneng-resync-audit.json")
SQL_PATH = Path("reports/makgabaneng-resync.sql")

TOTAL_TOKENS = {"total", "sub total", "subtotal", "total by gender"}
HEADER_TOKENS = {"age", "age sex", "age/sex"}


@dataclass
class IndicatorProxy:
    id: int
    code: str
    name: str
    is_active: bool = True


def run_psql(query: str) -> str:
    command = [
        "docker",
        "exec",
        "bonaso-local-stack-postgres-1",
        "psql",
        "-U",
        "bonaso",
        "-d",
        "bonaso",
        "-F",
        "\t",
        "-A",
        "-c",
        query,
    ]
    return subprocess.check_output(command, text=True)


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip(" .")


def normalize_token(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").lower().replace("_", " ").replace("-", " ")).strip()


def is_total_token(value) -> bool:
    return normalize_token(value) in TOTAL_TOKENS


def is_header_token(value) -> bool:
    return normalize_token(value) in HEADER_TOKENS


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def number_to_json(value):
    if value is None:
        return None
    if float(value).is_integer():
        return int(value)
    return float(value)


def first_numeric(values):
    for value in values:
        parsed = parse_number(value)
        if parsed is not None:
            return parsed
    return None


def canonical_sex(value):
    token = normalize_token(value)
    if token in {"m", "male"}:
        return "Male"
    if token in {"f", "female"}:
        return "Female"
    return None


def canonical_primary_label(value):
    text = clean_text(value)
    if not text:
        return "All"
    mapping = {
        "blood pressure messages": "Blood Pressure messages",
        "blood glucose messages": "Blood glucose messages",
        "tobacco control mssages": "Tobacco Control mssages",
        "living with a person with mental illness": "Living with a person with mental illness",
        "sleep hygiene": "Sleep hygiene",
    }
    return mapping.get(normalize_token(text), text)


def is_code_token(value) -> bool:
    if value is None:
        return False
    text = clean_text(value)
    return bool(re.fullmatch(r"\d+[a-z]?", text, flags=re.IGNORECASE))


def extract_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return clean_text(value)


def get_age_columns(header_row):
    age_columns = []
    total_col = None
    for index in range(6, len(header_row)):
        token = clean_text(header_row[index])
        normalized = normalize_token(token)
        if normalized == "total":
            total_col = index
            break
        if token:
            age_columns.append((index, token))
    return age_columns, total_col


def build_block_ranges(rows):
    code_rows = []
    for index, row in enumerate(rows):
        if is_code_token(row[1] if len(row) > 1 else None):
            code_rows.append(index)

    blocks = []
    for position, start in enumerate(code_rows):
        end = code_rows[position + 1] if position + 1 < len(code_rows) else len(rows)
        blocks.append((start, end))
    return blocks


def extract_block_title(rows, start, end):
    for row in rows[start : min(end, start + 4)]:
        title = clean_text(row[2] if len(row) > 2 else "")
        if title:
            return title
    return ""


def build_matrix_payload(detail_rows, header_row):
    age_columns, total_col = get_age_columns(header_row)
    if not age_columns:
        return None

    disaggregates = {}
    current_primary = "All"
    male_total = 0
    female_total = 0
    total = 0
    saw_sex = False
    saw_any_values = False
    final_total = None
    final_male_total = None
    final_female_total = None
    previous_row_was_total = False

    for row in detail_rows:
        row_title = clean_text(row[2] if len(row) > 2 else "")
        primary_label = clean_text(row[4] if len(row) > 4 else "")
        raw_sex = clean_text(row[5] if len(row) > 5 else "")
        sex = canonical_sex(raw_sex)
        numeric_age_values = [
            parse_number(row[column_index] if len(row) > column_index else None)
            for column_index, _ in age_columns
        ]
        numeric_age_value_count = sum(value is not None for value in numeric_age_values)
        first_age_value = numeric_age_values[0] if numeric_age_values else None

        if is_total_token(row_title) or is_total_token(primary_label) or is_total_token(raw_sex):
            numeric_total = first_numeric(row[6:])
            if sex == "Male":
                final_male_total = numeric_total
            elif sex == "Female":
                final_female_total = numeric_total
            elif numeric_total is not None:
                final_total = numeric_total
            previous_row_was_total = True
            continue

        if previous_row_was_total and not primary_label and numeric_age_value_count == 1:
            if sex == "Male":
                final_male_total = first_age_value
            elif sex == "Female":
                final_female_total = first_age_value
            elif final_total is not None and first_age_value is not None:
                final_total += first_age_value
            previous_row_was_total = False
            continue

        if primary_label:
            current_primary = canonical_primary_label(primary_label)

        age_values = {}
        row_sum = 0
        for column_index, age_band in age_columns:
            value = parse_number(row[column_index] if len(row) > column_index else None)
            if value is None:
                continue
            age_values[age_band] = number_to_json(value)
            row_sum += value

        if total_col is not None:
            explicit_total = parse_number(row[total_col] if len(row) > total_col else None)
            if explicit_total is not None:
                row_sum = explicit_total

        if not age_values and row_sum == 0:
            continue

        saw_any_values = True
        bucket_primary = current_primary if current_primary else "All"
        bucket_secondary = sex or "All"
        disaggregates.setdefault(bucket_primary, {})
        disaggregates[bucket_primary].setdefault(
            bucket_secondary,
            {band: 0 for _, band in age_columns},
        )
        for _, age_band in age_columns:
            disaggregates[bucket_primary][bucket_secondary][age_band] = (
                disaggregates[bucket_primary][bucket_secondary].get(age_band, 0)
                + age_values.get(age_band, 0)
            )

        total += row_sum
        if sex == "Male":
            saw_sex = True
            male_total += row_sum
        elif sex == "Female":
            saw_sex = True
            female_total += row_sum

        previous_row_was_total = False

    if not saw_any_values:
        return None

    if final_total is not None:
        total = final_total
    if final_male_total is not None:
        male_total = final_male_total
    if final_female_total is not None:
        female_total = final_female_total

    payload = {"total": number_to_json(total)}
    if disaggregates:
        payload["disaggregates"] = disaggregates
    if saw_sex or final_male_total is not None or final_female_total is not None:
        payload["male"] = number_to_json(male_total)
        payload["female"] = number_to_json(female_total)
    return payload


def build_simple_payload(block_rows):
    disaggregates = {}
    current_primary = "All"
    male_total = 0
    female_total = 0
    total = 0
    saw_sex = False
    saw_values = False
    final_total = None
    final_male_total = None
    final_female_total = None
    implicit_age_bands = [
        "10 -14",
        "15 -19",
        "20 - 24",
        "25 -29",
        "30 -34",
        "35 - 39",
        "40 -44",
        "45 -49",
        "50-54",
        "55-59",
        "60 -64",
        "65+",
    ]

    for row in block_rows:
        primary_label = clean_text(row[4] if len(row) > 4 else "")
        raw_sex = clean_text(row[5] if len(row) > 5 else "")
        sex = canonical_sex(raw_sex)
        age_slice = [parse_number(row[index] if len(row) > index else None) for index in range(6, 18)]
        explicit_total = parse_number(row[18] if len(row) > 18 else None)
        age_value_count = sum(value is not None for value in age_slice)
        numeric_total = first_numeric(row[6:])

        if is_total_token(primary_label) or is_total_token(raw_sex):
            if sex == "Male":
                final_male_total = numeric_total
            elif sex == "Female":
                final_female_total = numeric_total
            elif numeric_total is not None:
                final_total = numeric_total
            continue

        if primary_label:
            current_primary = canonical_primary_label(primary_label)

        if sex and age_value_count >= 2:
            age_values = {
                band: number_to_json(value)
                for band, value in zip(implicit_age_bands, age_slice)
                if value is not None
            }
            row_sum = explicit_total if explicit_total is not None else sum(value or 0 for value in age_slice)
            saw_values = True
            bucket_primary = current_primary if current_primary else "All"
            bucket_secondary = sex
            disaggregates.setdefault(bucket_primary, {})
            disaggregates[bucket_primary].setdefault(
                bucket_secondary,
                {band: 0 for band in implicit_age_bands},
            )
            for band in implicit_age_bands:
                disaggregates[bucket_primary][bucket_secondary][band] = (
                    disaggregates[bucket_primary][bucket_secondary].get(band, 0)
                    + age_values.get(band, 0)
                )

            total += row_sum
            saw_sex = True
            if sex == "Male":
                male_total += row_sum
            elif sex == "Female":
                female_total += row_sum
            continue

        if numeric_total is None:
            continue

        saw_values = True
        bucket_primary = current_primary if current_primary else "All"
        bucket_secondary = sex or "All"
        disaggregates.setdefault(bucket_primary, {})
        disaggregates[bucket_primary].setdefault(bucket_secondary, {"Value": 0})
        disaggregates[bucket_primary][bucket_secondary]["Value"] = number_to_json(
            (parse_number(disaggregates[bucket_primary][bucket_secondary].get("Value")) or 0)
            + numeric_total
        )

        total += numeric_total
        if sex == "Male":
            saw_sex = True
            male_total += numeric_total
        elif sex == "Female":
            saw_sex = True
            female_total += numeric_total

    if not saw_values:
        direct_total = first_numeric(block_rows[0][6:])
        if direct_total is None:
            return None
        return {"total": number_to_json(direct_total)}

    if final_total is not None:
        total = final_total
    if final_male_total is not None:
        male_total = final_male_total
    if final_female_total is not None:
        female_total = final_female_total

    payload = {"total": number_to_json(total)}
    if disaggregates:
        payload["disaggregates"] = disaggregates
    if saw_sex or final_male_total is not None or final_female_total is not None:
        payload["male"] = number_to_json(male_total)
        payload["female"] = number_to_json(female_total)
    return payload


def parse_sheet(ws):
    rows = [list(row) for row in ws.iter_rows(values_only=True, max_col=23)]
    parsed = []

    for start, end in build_block_ranges(rows):
        block_rows = rows[start:end]
        code = extract_code(rows[start][1] if len(rows[start]) > 1 else "")
        title = extract_block_title(rows, start, end)
        if not code or not title:
            continue

        header_index = None
        for index in range(max(0, start - 2), min(end, start + 6)):
            row = rows[index]
            if is_header_token(row[5] if len(row) > 5 else ""):
                header_index = index
                break

        if header_index is not None:
            detail_start = start
            header_row = rows[header_index]
            value = build_matrix_payload(rows[detail_start:end], header_row)
        else:
            value = build_simple_payload(block_rows)
        if value is None:
            continue

        dedupe_key = (code.lower(), canonical_indicator_key(title))
        parsed.append(
            {
                "dedupe_key": dedupe_key,
                "code": code,
                "title": title,
                "value": value,
            }
        )

    deduped = {}
    for item in parsed:
        deduped[item["dedupe_key"]] = item
    return list(deduped.values())


def load_indicators():
    query = """
select id, code, name, is_active
from indicators_indicator
order by id;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    _, *rows = lines

    indicator_candidates = {}
    for row in rows:
        indicator_id, code, name, is_active = row.split("\t", 3)
        candidate = IndicatorProxy(
            id=int(indicator_id),
            code=code,
            name=name,
            is_active=is_active == "t",
        )
        key = canonical_indicator_key(candidate.name)
        existing = indicator_candidates.get(key)
        if existing is None or get_indicator_resolution_priority(candidate) < get_indicator_resolution_priority(existing):
            indicator_candidates[key] = candidate
    return indicator_candidates


def load_organizations():
    query = """
with recursive orgs as (
  select id,name
  from organizations_organization
  where id = 5
  union all
  select o.id,o.name
  from organizations_organization o
  join orgs on o.parent_id = orgs.id
)
select id, name
from orgs
order by id;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    _, *rows = lines
    result = {}
    for row in rows:
        org_id, name = row.split("\t", 1)
        result[name] = int(org_id)
    return result


def load_existing_rows():
    query = f"""
with recursive orgs as (
  select id,name
  from organizations_organization
  where id = 5
  union all
  select o.id,o.name
  from organizations_organization o
  join orgs on o.parent_id = orgs.id
)
select
  a.id,
  o.name as organization_name,
  a.organization_id,
  a.indicator_id,
  i.code,
  i.name,
  a.value::text
from aggregates_aggregate a
join orgs o on o.id = a.organization_id
join indicators_indicator i on i.id = a.indicator_id
where a.project_id = {PROJECT_ID}
  and a.period_start = '{PERIOD_START}'
  and a.period_end = '{PERIOD_END}'
order by o.name, i.name;
""".strip()
    lines = [line for line in run_psql(query).splitlines() if line.strip() and "\t" in line]
    _, *rows = lines
    parsed = []
    for row in rows:
        aggregate_id, org_name, organization_id, indicator_id, code, name, value = row.split("\t", 6)
        parsed.append(
            {
                "aggregate_id": int(aggregate_id),
                "organization_name": org_name,
                "organization_id": int(organization_id),
                "indicator_id": int(indicator_id),
                "code": code,
                "name": name,
                "value": json.loads(value),
            }
        )
    return parsed


def main():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    indicator_candidates = load_indicators()
    organizations = load_organizations()
    existing_rows = load_existing_rows()
    existing_by_org_indicator = {
        (row["organization_name"], row["indicator_id"]): row
        for row in existing_rows
    }

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)

    updates = []
    inserts = []
    missing = []
    desired_keys = set()

    for sheet_name, org_name in SHEET_TO_ORG.items():
        org_id = organizations.get(org_name)
        if org_id is None:
            missing.append({"sheet": sheet_name, "organization": org_name, "reason": "organization missing"})
            continue

        items = parse_sheet(workbook[sheet_name])
        for item in items:
            indicator = resolve_indicator(item["title"], indicator_candidates, code=item["code"])
            if indicator is None:
                missing.append({"sheet": sheet_name, "organization": org_name, "code": item["code"], "title": item["title"], "reason": "indicator missing"})
                continue

            desired_keys.add((org_name, indicator.id))
            existing = existing_by_org_indicator.get((org_name, indicator.id))
            if existing is None:
                inserts.append(
                    {
                        "organization_name": org_name,
                        "organization_id": org_id,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "value": item["value"],
                    }
                )
                continue

            if existing["value"] != item["value"]:
                updates.append(
                    {
                        "aggregate_id": existing["aggregate_id"],
                        "organization_name": org_name,
                        "indicator_id": indicator.id,
                        "indicator_name": indicator.name,
                        "code": item["code"],
                        "title": item["title"],
                        "before": existing["value"],
                        "after": item["value"],
                    }
                )

    stale_rows = [
        row
        for row in existing_rows
        if (row["organization_name"], row["indicator_id"]) not in desired_keys
    ]

    sql_lines = ["begin;"]
    for update in updates:
        value_json = json.dumps(update["after"], ensure_ascii=True)
        sql_lines.append(
            "update aggregates_aggregate "
            f"set value = {sql_literal(value_json)}::jsonb, "
            "status = 'approved', "
            "reviewed_at = now(), "
            "notes = coalesce(notes,'') || ' | resynced from Makgabaneng workbook' "
            f"where id = {update['aggregate_id']};"
        )
    for insert in inserts:
        value_json = json.dumps(insert["value"], ensure_ascii=True)
        notes = (
            f"Imported from {WORKBOOK_PATH.name} | sheet={insert['organization_name']} | "
            f"code={insert['code']} | resynced from Makgabaneng workbook"
        )
        sql_lines.append(
            "insert into aggregates_aggregate "
            "(period_start, period_end, value, notes, created_at, updated_at, created_by_id, "
            "indicator_id, organization_id, project_id, reviewed_at, reviewed_by_id, status) values ("
            f"{sql_literal(PERIOD_START)}, "
            f"{sql_literal(PERIOD_END)}, "
            f"{sql_literal(value_json)}::jsonb, "
            f"{sql_literal(notes)}, "
            "now(), now(), null, "
            f"{insert['indicator_id']}, "
            f"{insert['organization_id']}, "
            f"{PROJECT_ID}, "
            "now(), null, 'approved');"
        )
    for row in stale_rows:
        sql_lines.append(
            f"delete from aggregates_aggregatechangelog where aggregate_id = {row['aggregate_id']};"
        )
        sql_lines.append(
            f"delete from aggregates_aggregate where id = {row['aggregate_id']};"
        )
    sql_lines.append("commit;")
    SQL_PATH.write_text("\n".join(sql_lines) + "\n", encoding="utf-8")

    report = {
        "workbook": str(WORKBOOK_PATH),
        "existing_rows": len(existing_rows),
        "updates": len(updates),
        "inserts": len(inserts),
        "stale_rows": len(stale_rows),
        "missing": missing,
        "updated_rows": updates[:200],
        "inserted_rows": inserts,
        "stale_row_details": stale_rows,
        "sql_path": str(SQL_PATH),
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "updates": len(updates),
                "inserts": len(inserts),
                "stale_rows": len(stale_rows),
                "missing": len(missing),
            },
            indent=2,
        )
    )
    print(f"Report: {REPORT_PATH}")
    print(f"SQL: {SQL_PATH}")


if __name__ == "__main__":
    main()
